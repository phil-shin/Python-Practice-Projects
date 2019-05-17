#! python
# spreadsheet2text.py
# Copies rows within each column of spreadsheet into separate .txt files

import openpyxl, os

# Set working directory
cwd = os.path.join('c:', os.sep, 'Users', 'Phil', 'Documents', 'python')
print(cwd)
os.chdir(cwd)

# Open spreadsheet
wb=openpyxl.load_workbook('text2spreadsheet.xlsx')
sheet=wb.active
# Write row contents in each column to separate .txt files
for col in range(1, sheet.max_column+1):
    fileName='ss2text'+str(col)+'.txt'
    file=open(fileName, 'a')
    for row in range(1, sheet.max_row+1):
        file.write(sheet.cell(row=row, column=col).value)
    file.close()

