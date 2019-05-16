#! python
# cellInverter.py
# Inverts cells of a spreadsheet

import openpyxl, sys, os

# Change directory
cwd = os.path.join('c:', os.sep, 'Users','PS81581','AppData','Local','Programs','Python','Python37')
os.chdir(cwd)

# Open spreadsheet and read contents
wb = openpyxl.load_workbook('cellInvert.xlsx')
sheet=wb.active

# Initialize list of lists data structure
sheetData = [[row]*sheet.max_column for row in range(1, sheet.max_row+1)]
#print(sheetData)

# Loop spreadsheet content into data structure
for row in range(1, sheet.max_row+1):
    for col in range(1, sheet.max_column+1):
        #print('row:')
        #print(row)
        #print('col:')
        #print(col)
        #print(sheet.cell(row=row, column=col).value)
        sheetData[row-1][col-1] = sheet.cell(row=row, column=col).value
#print(sheetData)

# Create new spreadsheet and sheet
wbInvert= openpyxl.Workbook()
sheetInvert = wbInvert.active

# Loop to write inverted content into new spreadsheet
for row in range(1, sheet.max_column+1):
    for col in range(1, sheet.max_row+1):
        #print('row:')
        #print(row)
        #print('col:')
        #print(col)
        #print(sheetData[col-1][row-1])
        sheetInvert.cell(row=row, column=col).value=sheetData[col-1][row-1]

# Save spreadsheet
wbInvert.save('cellInvertNew.xlsx')
