#! python
# blankRowInserter.py
# Takes command line arguments N and M and inserts M rows following row N

import openpyxl, sys, os

# Change directory
cwd = os.path.join('c:', os.sep, 'Users','PS81581','AppData','Local','Programs','Python','Python37')
os.chdir(cwd)

# Store command line arguments
N = int(sys.argv[1])
M = int(sys.argv[2])

# Open spreadsheet and read contents
wb = openpyxl.load_workbook('blankrow.xlsx')
sheet=wb.active

# Open new blank row spreadsheet
wbBlank = openpyxl.Workbook()
sheetBlank = wbBlank.active

# Rewrite first N rows in new spreadsheet
for row in range(1, N+1):
    for col in range(1, sheet.max_column+1):
        cellValue = sheet.cell(row=row, column=col).value
        sheetBlank.cell(row=row, column=col).value=cellValue

# Insert M blank rows
for row in range(N+1, N+M+1):
    sheetBlank.cell(row=row, column=1).value=''
    
# Rewrite remaining rows in new spreadsheet
for row in range(N+M+1, sheet.max_row+1+M):
    for col in range(1, sheet.max_column+1):
        cellValue = sheet.cell(row=row-M, column=col).value
        sheetBlank.cell(row=row, column=col).value=cellValue
        
# Save and Close Spreadsheet
wbBlank.save('blankrowNew.xlsx')
