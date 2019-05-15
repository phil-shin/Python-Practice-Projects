#! python
# multiplicationTable.py
# Takes command line N dimension input and creates an NxN multiplcation table

import sys, openpyxl, os
from openpyxl.styles import Font
wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb['Sheet']


#Prompt user for dimension
#print('Please enter dimension for square multiplication table:')
n = int(sys.argv[1])
print(n)

# Create Table structure
boldFont = Font(bold=True)
num=1
for i in range(2, n+2):
    #print(i)
    sheet.cell(row=1, column=i).value = num
    sheet.cell(row=1, column=i).font=boldFont
    sheet.cell(row=i, column=1).value = num
    sheet.cell(row=i, column=1).font=boldFont
    num+=1
    
#row1 = sheet[1]
#print(row1)
#print(row1[0])
#print(row1[1])
#colA = sheet['A']
#print(colA)
#for i in range(0,n+1):
    #print(i)
    #row1[i].value = i
    #row1[i].font=boldFont
    #colA[i].value = i
    #colA[i].font = boldFont
 
# Loop through cells to multiply
num2=1
for row in range(2, sheet.max_row+1):
    num3=1
    for col in range(2, sheet.max_column+1):
        sheet.cell(row=row, column = col).value = '='+str(num2)+'*'+str(num3)
        num3+=1
    num2+=1
        
# Save excel spreadsheet
filepath = os.path.join('c:', os.sep, 'Users','PS81581','AppData','Local','Programs','Python','Python37')
os.chdir(filepath)
wb.save('multTable.xlsx')
